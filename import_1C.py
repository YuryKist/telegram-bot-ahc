# import_1C.py
"""
Модуль для обработки выгрузки из 1С и обновления Реестра АХЧ.
Ожидает:
- Файл *Реестр АХЧ*.xlsx — основной реестр
- Файл *Платежн*.xlsx — выписка из 1С
Обновляет статус оплаты в реестре и возвращает результат.
"""

import pandas as pd
import re
from pathlib import Path
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell import MergedCell


# Настройка логирования
logger.remove()
logger.add("1C_import.log", rotation="10 MB", level="INFO", encoding="utf-8")
logger.add(lambda msg: print(msg, end=''), level="INFO")



def load_ahx_data(directory_path):
    """Загружает реестр из Excel (начиная с 4 строки)."""
    try:
        directory =  Path(directory_path)
        all_excel_files = list(directory.glob('*.xls*'))
        excel_files_AXCH = [
            f for f in all_excel_files 
            if 'ахч' in f.name.lower()
            ]
        if not excel_files_AXCH:
            raise FileNotFoundError("❌ Файл выписки из АХЧ не найден.")
        
        ahch_file_path = excel_files_AXCH[0]
        
        df_register = pd.read_excel(ahch_file_path, skiprows=3)
        logger.info(f"📊 Загружено {len(df_register)} записей из реестра.")
        return df_register, ahch_file_path
    
    except Exception as e:
        logger.error(f"[Ошибка загрузки Excel] {e}")
        return None

def load_payment_data(directory_path):
    """Загружает выписку из 1С."""
    try:
        directory = Path(directory_path)
        excel_files_payment = list(directory.glob('*Платежн*.xls*'))

        if not excel_files_payment:
            raise FileNotFoundError("❌ Файл выписки из 1С не найден.")

        df_export_1C = pd.read_excel(excel_files_payment[0], skiprows=4)
        logger.info(f"📊 Загружено {len(df_export_1C)} записей из реестра.")
        return df_export_1C
    
    except Exception as e:
        logger.error(f"❌ Ошибка при загрузке файлов: {e}")
        return None


def extract_invoice_number(text):
    """Извлекает номер документа из текста (счёт, накладная и т.д.)."""
    if not isinstance(text, str):
        return None

    invoice_chars = r"[\w\-_+\/A-Za-zА-Яа-яЁё]"

    # 1. Ищем после ключевых слов: "счет", "накладная" и т.д.
    main_pattern = rf"""
        (?:счет[ау]?|фактур[еа]|накладн[оийя]|товарн[аы][йя]|тмт|с\/?ф|[сc]\/?ф|тов\.?\s*накладн[оийя])
        \b                      # граница слова
        \W*                     # разделители
        ({invoice_chars}*?\d{invoice_chars}*)  # номер с цифрой
    """
    match = re.search(main_pattern, text, re.IGNORECASE | re.DOTALL | re.VERBOSE)
    if match:
        return match.group(1).strip()

    # 2. Ищем после символа №
    alt_match = re.search(r"№\s*([A-Za-zА-Яа-яЁё\d\-_+\/]+)", text, re.IGNORECASE)
    if alt_match:
        return alt_match.group(1).strip()

    # 3. Берём первое число
    fallback_match = re.search(r"\b\d+\b", text)
    if fallback_match:
        return fallback_match.group(0).strip()

    logger.debug("Номер документа не найден.")
    return None

def prepare_register(df_register):
    """Подготавливает реестр: убирает ведущие нули, приводит типы."""
    logger.info("Преобразование типов данных реест АХЧ")
    df_register = df_register.copy()
    df_register['№ счета'] = df_register['№ счета'].fillna('').astype("string")
    df_register['Сумма'] = pd.to_numeric(df_register['Сумма'], errors='coerce').round(2)
    df_register['№ задачи Битрикс'] = pd.to_numeric(df_register['№ задачи Битрикс'], errors='coerce').astype(dtype = int, errors = 'ignore')
    df_register['ID_Счет_Bitrix'] = pd.to_numeric(df_register['ID_Счет_Bitrix'], errors='coerce').astype(dtype = int, errors = 'ignore')

    logger.info("✅ Данные подготовлены.")
    return df_register


def prepare_export_1C(df_export_1C):
    """Подготавливает данные: убирает ведущие нули, приводит типы."""
    
    logger.info("Преобразование типов данных реестра 1C")
    df_export_1C = df_export_1C.copy()
    # Извлекаем номера счетов из "Информация"
    df_export_1C['Номер счета'] = df_export_1C['Информация'].apply(extract_invoice_number)
    df_export_1C['Номер счета'] = df_export_1C['Номер счета'].astype(str).apply(
        lambda x: str(int(x)) if x.isdigit() else x
    )

    # Приводим суммы к числу
    df_export_1C['Сумма'] = pd.to_numeric(df_export_1C['Сумма'], errors='coerce').round(2)

    logger.info("✅ Данные подготовлены.")
    return df_export_1C


def update_payment_status(df_export, df_register):
    """Обновляет статус оплаты на основе совпадения по № счёта и сумме."""
    try:
        # Убираем дубликаты в выписке
        df_export_unique = df_export.drop_duplicates(subset=['Номер счета', 'Сумма'])
        df_export_mapped = df_export_unique[['Номер счета', 'Сумма', 'Состояние']].copy()
        df_export_mapped.columns = ['№ счета', 'Сумма', 'Статус оплаты']

        # Объединяем
        df_merged = df_register.merge(
            df_export_mapped,
            on=['№ счета', 'Сумма'],
            how='left',
            suffixes=('', '_новый')
        )

        # Обновляем только пустые или неокончательные статусы
        mask = df_merged['Статус оплаты_новый'].notna() & \
               (~df_register['Статус оплаты'].isin(["Оплачено", "Подготовлено"]))
        df_register = df_register.copy()
        df_register.loc[mask, 'Статус оплаты'] = df_merged.loc[mask, 'Статус оплаты_новый']

        updated_count = mask.sum()
        if updated_count > 0:
            logger.info(f"✅ Обновлено {updated_count} статусов оплаты.")
        else:
            logger.info("ℹ️ Новых статусов для обновления не найдено.")

        return df_register
    except Exception as e:
        logger.error(f"❌ Ошибка при обновлении статуса: {e}")
        return df_register


def update_date_payment(df_register):
    # Словарь с количеством дней для каждого поставщика
    days_dict = {
        'ип шайдулин': 30,
        'технология 21век': 0,
        'корексмаркет': 28,
        'регион-снабжение': 28,
        'курышев': 25,
        'упаковочные материалы': 30,
        'ип павлов е.в.': 0
    }
    
    # Создаем копию датафрейма для безопасной работы
    df_result = df_register.copy()
    
    # Инициализируем счетчик обновленных записей
    updated_count = 0
    
    try:
        # Преобразуем колонку 'Дата счета' в datetime с указанием формата
        df_result['Дата счета'] = pd.to_datetime(df_result['Дата счета'], errors='coerce', format='%d.%m.%Y', dayfirst=True)
        
        # Создаем или очищаем колонку 'Контроль оплаты' если она не существует
        if 'Контроль оплаты' not in df_result.columns:
            df_result['Контроль оплаты'] = None
        
        # Проходим по каждой строке датафрейма
        for index, row in df_result.iterrows():
            try:
                # Проверяем, что дата счета не пустая
                if pd.notna(row['Дата счета']):
                    # Получаем поставщика и приводим к нижнему регистру
                    supplier = str(row['Поставщик']).lower().strip()
                    
                    # Получаем количество дней из словаря
                    days_to_add = days_dict.get(supplier, 0)
                    
                    # Добавляем дни к дате счета
                    new_date = row['Дата счета'] + pd.Timedelta(days=days_to_add)
                    # Записываем в колонку 'Контроль оплаты'
                    df_result.at[index, 'Контроль оплаты'] = new_date
                    # Увеличиваем счетчик обновлений
                    updated_count += 1
                    
            except Exception as e:
                logger.error(f"Ошибка при обработке строки {index}: {e}")
                continue
        
        # Преобразуем колонку 'Контроль оплаты' в строковый формат
        if 'Контроль оплаты' in df_result.columns and not df_result['Контроль оплаты'].empty:
            df_result['Контроль оплаты'] = pd.to_datetime(df_result['Контроль оплаты'], errors='coerce').dt.strftime('%d.%m.%Y')
        
        # Логируем результат
        logger.info(f"✅ Обновлено {updated_count} дат контроля оплаты.")
        
    except Exception as e:
        logger.error(f"Общая ошибка при обновлении дат: {e}")
        logger.error(f"Тип данных в колонке 'Дата счета': {df_result['Дата счета'].dtype if 'Дата счета' in df_result.columns else 'Колонка отсутствует'}")
    
    return df_result


def update_excel_file(file_path, df):
    """Сохраняет DataFrame в Excel с сохранением форматирования."""
    try:
        df_save = df.copy()

        # приводим формат дат к EXCEL
        date_columns = ['Контроль оплаты', 'Дата счета']
        for col in date_columns:
            if col in df_save.columns:
                if pd.api.types.is_datetime64_any_dtype(df_save[col]):
                    df_save[col] = df_save[col].dt.strftime('%d.%m.%Y')
                elif df_save[col].dtype == 'object':
                    try:
                        parsed = pd.to_datetime(
                            df_save[col], 
                            errors='coerce',
                            dayfirst=True,
                            format='%d.%m.%Y'
                        )
                        if parsed.notna().any():
                            df_save[col] = parsed.dt.strftime('%d.%m.%Y').where(parsed.notna(), df_save[col])
                    except Exception as e:
                        print(f"Ошибка при обработке колонки '{col}': {e}")
                        pass

        wb = load_workbook(file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None

        for col_idx, col_name in enumerate(df_save.columns, 1):
            ws.cell(row=4, column=col_idx, value=col_name)

        for row_idx, row in enumerate(dataframe_to_rows(df_save, index=False, header=False), 5):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)
        logger.info(f"💾 Реестр обновлён: {file_path}")
        return True
    except Exception as e:
        logger.error(f"❌ Ошибка при сохранении Excel: {e}")
        return False


def run_pipeline(directory_path: str) -> str:
    """
    Добавляет данные из выписки из 1С в реестр.
    Возвращает строку-результат для Telegram-бота.
    """
    try:
        logger.info("🔹 Запуск пайплайна: обработка выписки из 1С")

        # Шаг 1: Загрузка файлов
        df_register, ahch_path = load_ahx_data(directory_path)
        df_export_1C = load_payment_data(directory_path)

        # Шаг 2: Подготовка данных
        df_export_clean = prepare_export_1C(df_export_1C)
        df_register_clean = prepare_register(df_register)

        # Шаг 3: Обновление данных
        df_register_updadate = update_payment_status(df_export_clean, df_register_clean)
        df_register_updadate = update_date_payment(df_register_updadate)

        # Шаг 4: Сохранение
        success = update_excel_file(ahch_path, df_register_updadate)
        if success:
            return "✅ Данные из 1С обновлены в Реестре АХЧ."
        else:
            return "❌ Ошибка при сохранении файла."

    except Exception as e:
        logger.error(f"❌ Критическая ошибка: {e}")
        return f"❌ Ошибка выполнения: {str(e)}"


# --- Для теста ---
if __name__ == "__main__":
    test_path = r"C:\Users\Юрий Кистенев\Desktop\ACH_manager\record"
    result = run_pipeline(test_path)
    print(result)